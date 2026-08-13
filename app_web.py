import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import hashlib
import io
import plotly.graph_objects as go
import plotly.express as px
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.parse
import pytz
import re
import os
import base64
import requests

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from funcoes import (
    buscar_todas_movimentacoes,
    dados_dashboard,
    dados_grafico_mensal,
    dados_grafico_categorias,
    alterar_senha_usuario,
    obter_limite_orcamento,
    definir_limite_orcamento,
    cadastrar_conta,
    listar_contas,
    excluir_conta,
    criar_meta,
    listar_metas,
    atualizar_progresso_meta,
    salvar_movimentacao,
    salvar_movimentacao_parcelada,
    salvar_movimentacao_recorrente,
    excluir_movimentacao,
    formatar_moeda_ptbr,
    listar_todos_usuarios_admin,
    atualizar_status_e_mensalidade,
    excluir_usuario_admin,
    excluir_meta,
    obter_limites_por_categoria,
    buscar_vencimentos_proximos,
    salvar_orcamento_categoria,
    excluir_orcamento_categoria,
    excluir_lancamento_pendente,
    marcar_lancamento_como_pago,
    desfazer_pagamento_lancamento,
    dar_baixa_fatura_completa,
    listar_cartoes,
    cadastrar_cartao,
    calcular_mes_fatura,
    buscar_gastos_fatura,
    atualizar_limite_cartao,
    excluir_cartao,
    dados_grafico_tags,
    gerar_insights_financeiros,
    excluir_conta_a_receber,
    alternar_status_contas_a_receber,
    salvar_conta_a_receber,
    buscar_contas_a_receber,
    atualizar_conta_a_receber,
    buscar_movimentacoes_paginadas,
    obter_caminho_logo
)

from views import render_sidebar_footer

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="FinanceiroPro Web",
    page_icon="logo",
    layout="wide",
    initial_sidebar_state="expanded",
)

url_uid = st.query_params.get("uid")

if "usuario_id" not in st.session_state:
    st.session_state["usuario_id"] = int(url_uid) if url_uid else None

st.session_state["is_admin"] = st.session_state.get("usuario_id") == 1
def fmt_moeda(valor):
    return formatar_moeda_ptbr(float(valor))

def criar_link_cobranca(telefone, nome, valor):
    if not telefone:
        return None
    telefone_limpo = "".join(filter(str.isdigit, str(telefone)))
    mensagem = (
        f"Olá, {nome}! Tudo bem?\n\n"
        f"Passando para lembrar que a mensalidade da sua plataforma FinanceiroPro "
        f"no valor de R$ {valor:.2f} está em aberto.\n\n"
        f"Para restabelecer ou manter o seu acesso integral, você pode realizar o pagamento. "
        f"Caso já tenha efetuado, por favor, envie o comprovante por aqui. Obrigado!"
    )
    texto_codificado = urllib.parse.quote(mensagem)
    return f"https://api.whatsapp.com/send?phone=55{telefone_limpo}&text={texto_codificado}"

CATEGORIAS_DESPADREVAL = ["Alimentação", "Transporte", "Cartão de Crédito", "Cartão de Débito", "Pix", "Moradia", "Lazer", "Saúde", "Educação", "Assinaturas/Serviços", "Outros"]
CATEGORIAS_RECEITAS = ["Salário", "Freelance", "Investimentos", "Presente/Prêmio"]

def gerar_excel_profissional(dados_banco, mes, ano):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato Financeiro"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Relatório de Extrato Detalhado - Período: {mes}/{ano}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "Data", "Conta", "Tipo", "Forma Pagto", "Descrição", "Valor", "Categoria", "Tags"]
    ws.append([])
    ws.append(headers)
    ws.row_dimensions[3].height = 26

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    row_num = 4
    total_receitas = 0.0
    total_despesas = 0.0

    for item in dados_banco:
        v_id, v_data, v_conta, v_tipo, v_forma, v_desc, v_valor, v_cat = item[:8]
        v_tags = item[8] if len(item) > 8 else ""
        v_valor_float = float(v_valor)

        try:
            v_data_fmt = pd.to_datetime(v_data).strftime("%d/%m/%Y")
        except Exception:
            v_data_fmt = str(v_data)

        if str(v_tipo).lower() == "receita":
            total_receitas += v_valor_float
        else:
            total_despesas += v_valor_float

        ws.append([v_id, v_data_fmt, v_conta, v_tipo, v_forma, v_desc, v_valor_float, v_cat, v_tags])
        ws.row_dimensions[row_num].height = 20

        bg_color = "F9FBFD" if row_num % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        for col_num in range(1, 10):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)

            if col_num in [1, 2, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '"R$"#,##0.00'
                cell.font = Font(name="Arial", size=10, color="27AE60" if str(v_tipo).lower() == "receita" else "C0392B")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1

    ws.append([])
    row_num += 1

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell_lbl_rec = ws.cell(row=row_num, column=1, value="Total Receitas:")
    cell_lbl_rec.font = Font(name="Arial", size=10, bold=True)
    cell_lbl_rec.alignment = Alignment(horizontal="right", vertical="center")

    cell_val_rec = ws.cell(row=row_num, column=7, value=total_receitas)
    cell_val_rec.font = Font(name="Arial", size=10, bold=True, color="27AE60")
    cell_val_rec.number_format = '"R$"#,##0.00'
    cell_val_rec.alignment = Alignment(horizontal="right", vertical="center")

    row_num += 1

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell_lbl_des = ws.cell(row=row_num, column=1, value="Total Despesas:")
    cell_lbl_des.font = Font(name="Arial", size=10, bold=True)
    cell_lbl_des.alignment = Alignment(horizontal="right", vertical="center")

    cell_val_des = ws.cell(row=row_num, column=7, value=total_despesas)
    cell_val_des.font = Font(name="Arial", size=10, bold=True, color="C0392B")
    cell_val_des.number_format = '"R$"#,##0.00'
    cell_val_des.alignment = Alignment(horizontal="right", vertical="center")

    row_num += 1

    grey_fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell_lbl_saldo = ws.cell(row=row_num, column=1, value="SALDO LÍQUIDO:")
    cell_lbl_saldo.font = Font(name="Arial", size=10, bold=True)
    cell_lbl_saldo.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx in range(1, 8):
        ws.cell(row=row_num, column=col_idx).fill = grey_fill

    saldo_final = total_receitas - total_despesas
    cell_val_saldo = ws.cell(row=row_num, column=7, value=saldo_final)
    cell_val_saldo.font = Font(name="Arial", size=10, bold=True, color="000000")
    cell_val_saldo.number_format = '"R$"#,##0.00'
    cell_val_saldo.alignment = Alignment(horizontal="right", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def gerar_pdf_profissional(dados_banco, mes, ano):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18,
        textColor=colors.HexColor('#1F4E78'), spaceAfter=6, alignment=1
    )
    subtitle_style = ParagraphStyle(
        'SubTitleStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10,
        textColor=colors.HexColor('#7F8C8D'), spaceAfter=20, alignment=1
    )

    story.append(Paragraph("FinanceiroPro Web - Extrato Detalhado", title_style))
    story.append(Paragraph(f"Período consultado: {mes}/{ano} — Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    
    table_data = [["ID", "Data", "Conta", "Tipo", "Forma Pagto", "Descrição", "Valor", "Categoria", "Tags"]]
    total_rec, total_des = 0.0, 0.0
    
    for item in dados_banco:
        v_id, v_data, v_conta, v_tipo, v_forma, v_desc, v_valor, v_cat = item[:8]
        v_tags = item[8] if len(item) > 8 else ""
        v_valor_float = float(v_valor)
        
        try:
            v_data_fmt = pd.to_datetime(v_data).strftime('%d/%m/%Y')
        except Exception:
            v_data_fmt = str(v_data)

        if str(v_tipo).lower() == "receita":
            total_rec += v_valor_float
        else:
            total_des += v_valor_float
            
        table_data.append([
            str(v_id), v_data_fmt, str(v_conta), str(v_tipo), 
            str(v_forma), str(v_desc), formatar_moeda_ptbr(v_valor_float), str(v_cat), str(v_tags)
        ])
        
    table_data.append(["", "", "", "", "", "Total Receitas:", formatar_moeda_ptbr(total_rec), "", ""])
    table_data.append(["", "", "", "", "", "Total Despesas:", formatar_moeda_ptbr(total_des), "", ""])
    table_data.append(["", "", "", "", "", "SALDO LÍQUIDO:", formatar_moeda_ptbr(total_rec - total_des), "", ""])
    
    t = Table(table_data, colWidths=[25, 60, 80, 50, 75, 130, 80, 85, 85])
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -4), 0.5, colors.HexColor('#E2E8F0')),
    ])
    
    for i in range(1, len(dados_banco) + 1):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC'))
            
    total_idx = len(table_data)
    t_style.add('FONTNAME', (5, total_idx-3), (6, total_idx-1), 'Helvetica-Bold')
    t_style.add('TEXTCOLOR', (6, total_idx-3), (6, total_idx-3), colors.HexColor('#27AE60'))
    t_style.add('TEXTCOLOR', (6, total_idx-2), (6, total_idx-2), colors.HexColor('#C0392B'))
    t_style.add('BACKGROUND', (5, total_idx-1), (6, total_idx-1), colors.HexColor('#EAEDED'))
    
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def fazer_login_rest(usuario, senha):
    BASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
    headers = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
    senha_hash = hashlib.sha256(senha.encode('utf-8')).hexdigest()
    url = f"{BASE_URL}/usuarios?usuario=eq.{usuario.strip()}&select=id,senha,role,status"
    res = requests.get(url, headers=headers)
    if res.status_code == 200 and res.json():
        usuario_banco = res.json()[0]
        if usuario_banco['senha'] == senha_hash:
            status_atual = usuario_banco.get('status', 'ativo')
            if status_atual == 'inativo':
                st.error("⚠️ Sua conta está suspensa. Entre em contato com o administrador.")
                return None
            elif status_atual == 'pendente':
                st.warning("⏳ Seu acesso está em análise! Aguarde a liberação do administrador.")
                return None
                
            st.session_state["is_admin"] = (usuario_banco.get('role') == 'admin')
            return usuario_banco['id']
    return None

def criar_usuario_rest(usuario, senha, status='pendente', valor_mensalidade=0.0, telefone=""):
    BASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
    headers = {
        "apikey": SUPABASE_KEY, 
        "Authorization": f"Bearer {SUPABASE_KEY}", 
        "Content-Type": "application/json", 
        "Prefer": "return=representation"
    }
    
    url_check = f"{BASE_URL}/usuarios?usuario=eq.{usuario.strip()}&select=id"
    res_check = requests.get(url_check, headers=headers)
    if res_check.status_code == 200 and res_check.json():
        return "Existe"
        
    senha_hash = hashlib.sha256(senha.encode('utf-8')).hexdigest()
    url_ins = f"{BASE_URL}/usuarios"
    
    payload = {
        "usuario": usuario.strip(), 
        "senha": senha_hash, 
        "role": "user", 
        "status": status, 
        "valor_mensalidade": valor_mensalidade,
        "telefone": telefone.strip()
    }
    
    res_ins = requests.post(url_ins, headers=headers, json=payload)
    if res_ins.status_code in [200, 201] and res_ins.json():
        return res_ins.json()[0]['id']
    return False

# --- AUTENTICAÇÃO ---
if st.session_state.get("usuario_id") is None:
    st.markdown(
        """
        <style>
        [data-testid="stSidebar"], [data-testid="stHeader"], footer { display: none !important; }
        .stApp { background-color: #0b0f19; }
        .stTabs [data-baseweb="tab-list"] {
            gap: 4px; background-color: #0f141d; padding: 4px; border-radius: 10px; border: 1px solid #232d3f;
        }
        .stTabs [data-baseweb="tab"] {
            height: 36px; border-radius: 8px; color: #94a3b8; font-weight: 500; border: none !important; flex-grow: 1; justify-content: center;
        }
        .stTabs [aria-selected="true"] { background-color: #2563eb !important; color: #ffffff !important; }
        .stTextInput input { background-color: #0f141d !important; color: #ffffff !important; border: 1px solid #232d3f !important; border-radius: 8px !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

usuario_atual_id = st.session_state.get("usuario_id") or st.query_params.get(
    "uid"
)

# Se NÃO estiver logado, exibe a tela de login
if not usuario_atual_id:
    # Define as colunas com base no dispositivo (Mobile vs Desktop)
    if st.session_state.get("is_mobile", False):
        _, col_center, _ = st.columns([0.1, 0.8, 0.1])
    else:
        _, col_center, _ = st.columns([1, 1.2, 1])

    # Todo o formulário de login DEVE estar indentado dentro do IF
    with col_center:
        with st.container(border=True):
            caminho_logo = obter_caminho_logo()
            if caminho_logo:
                _, col_img, _ = st.columns([1, 2, 1])
                with col_img:
                    st.image(caminho_logo, width="stretch")

            st.markdown(
                "<p style='text-align: center; color: #94a3b8; font-size: 12px; margin-top: 5px; margin-bottom: 15px;'>"
                "Acesse sua conta para gerenciar suas finanças"
                "</p>",
                unsafe_allow_html=True,
            )

            # (Coloque aqui os st.text_input e st.button de login)

            tab_login, tab_cadastro = st.tabs(["🔑 Fazer Login", "📝 Criar Nova Conta"])

            with tab_login:
                with st.form("form_login", clear_on_submit=False):
                    username_input = st.text_input("Seu Usuário", key="login_user", placeholder="Digite seu usuário")
                    password_input = st.text_input("Senha", type="password", key="login_pass", placeholder="••••••••")
                    btn_login = st.form_submit_button("Entrar no Sistema", type="primary", width="stretch")

                if btn_login:
                    if not username_input.strip() or not password_input.strip():
                        st.error("Preencha o usuário e a senha para entrar.")
                    else:
                        uid = fazer_login_rest(username_input, password_input)
                        if uid:
                            st.session_state["usuario_id"] = uid
                            st.query_params["uid"] = str(uid)
                            st.success("Logado com sucesso!")
                            st.rerun()
                        else:
                            st.error("Usuário incorreto, senha inválida ou restrição de acesso.")

            with tab_cadastro:
                with st.form("form_cadastro", clear_on_submit=False):
                    username_cad = st.text_input("Escolha um Usuário", key="cad_user", placeholder="Ex: joao.silva")
                    telefone_cad = st.text_input("Telefone (WhatsApp)", key="cad_tel", placeholder="556599998888")
                    password_cad = st.text_input("Crie uma Senha", type="password", key="cad_pass", placeholder="••••••••")
                    btn_cad = st.form_submit_button("Cadastrar e Solicitar Acesso", type="primary", width="stretch")

                if btn_cad:
                    if username_cad.strip() and telefone_cad.strip() and password_cad.strip():
                        res = criar_usuario_rest(
                            usuario=username_cad,
                            senha=password_cad,
                            telefone=telefone_cad,
                            status="pendente",
                        )

                        if res == "Existe":
                            st.error("Este nome de usuário já está em uso.")
                        elif res:
                            st.success("🎉 Cadastro realizado! Aguarde a liberação do administrador.")
                        else:
                            st.error("Erro ao realizar o cadastro. Tente novamente.")
                    else:
                        st.error("Por favor, preencha todos os campos obrigatórios.")

    st.stop()

# --- NAVEGAÇÃO E BARRA LATERAL ---
st.markdown(
    """
    <style>
    [data-testid="stSidebar"] > div:first-child { padding-top: 0.5rem !important; padding-bottom: 0.5rem !important; }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0px !important; padding: 0 !important; }
    [data-testid="stSidebar"] [data-testid="stImage"] { display: flex !important; justify-content: center !important; align-items: center !important; width: 100% !important; margin: 0 auto !important; }
    [data-testid="stSidebar"] img { max-width: 145px !important; width: 145px !important; margin: 0 auto 4px auto !important; display: block !important; }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] div { font-size: 0.75rem !important; margin: 2px 0 !important; text-align: center !important; }
    [data-testid="stSidebar"] hr { margin: 4px 0 !important; }
    [data-testid="stSidebar"] .stButton button { height: 34px !important; min-height: 34px !important; padding: 0px 8px !important; font-size: 12.5px !important; font-weight: 500 !important; border-radius: 6px !important; margin: 1px 0 !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

if "pagina_atual" not in st.session_state:
    st.session_state["pagina_atual"] = "📊 Dashboard"

opcoes_menu = [
    "📊 Dashboard",
    "🏦 Gerir Contas",
    "💸 Lançar Movimentações",
    "💳 Cartões & Faturas",
    "💰 Contas a Receber",
    "📅 Próximos Vencimentos",
    "🎯 Metas de Economia", 
    "🎯 Orçamentos por Categoria",
    "📋 Extrato Detalhado", 
    "⚙️ Configurações"
]

if st.session_state.get("is_admin", False):
    opcoes_menu.append("👑 Painel Admin SaaS")

with st.sidebar:
    caminho_logo = obter_caminho_logo()
    if caminho_logo:
        st.image(caminho_logo, width="stretch")
        
    st.write(f"👤 Usuário: **{st.session_state.get('usuario_id', '')}** {'(👑 Admin)' if st.session_state.get('is_admin') else ''}")
    st.markdown("---")
    
    for item in opcoes_menu:
        eh_ativo = (st.session_state["pagina_atual"] == item)
        tipo_botao = "primary" if eh_ativo else "secondary"
        
        if st.button(item, key=f"btn_nav_{item}", type=tipo_botao, width="stretch"):
            st.session_state["pagina_atual"] = item
            st.rerun()
    
    st.markdown("---")
    if st.button("🚪 Sair do Sistema", width="stretch"):
        st.query_params.clear()
        if "usuario_id" in st.session_state:
            del st.session_state["usuario_id"]
        if "is_admin" in st.session_state:
            del st.session_state["is_admin"]
        if "pagina_atual" in st.session_state:
            del st.session_state["pagina_atual"]
        st.rerun()

    render_sidebar_footer()

opcao = st.session_state["pagina_atual"]

# --- PAINEL ADMIN SAAS ---
if opcao == "👑 Painel Admin SaaS" and st.session_state.get("is_admin"):
    st.title("👑 Painel de Controle Master SaaS")
    
    usuarios_lista = listar_todos_usuarios_admin()
    df_users = pd.DataFrame(usuarios_lista)
    
    faturamento_mrr = 0.0
    if not df_users.empty and 'valor_mensalidade' in df_users.columns:
        faturamento_mrr = df_users[df_users['status'] == 'ativo']['valor_mensalidade'].astype(float).sum()
        
    st.metric("💰 Faturamento Mensal Estimado (Ativos)", fmt_moeda(faturamento_mrr))
    
    if not df_users.empty:
        col_fat1, col_fat2 = st.columns(2)
        col_fat1.metric("👥 Total Clientes", len(df_users))
        col_fat2.metric("⏳ Pendentes", len(df_users[df_users['status'] == 'pendente']))

    st.markdown("---")
    adm_tab1, adm_tab2, adm_tab3 = st.tabs(["⏳ Liberar Cadastros", "➕ Criar Cliente Manual", "👥 Gerenciar Clientes"])
    
    with adm_tab1:
        col_t1, col_t2 = st.columns([4, 1])
        with col_t1:
            st.write("### 🔑 Clientes aguardando liberação")
        with col_t2:
            if st.button("🔄 Atualizar", key="btn_ref_pend"):
                st.rerun()

        if not df_users.empty and 'status' in df_users.columns:
            df_pendentes = df_users[df_users['status'] == 'pendente']
            if df_pendentes.empty:
                st.success("Nenhum cliente aguardando liberação no momento.")
            else:
                for idx, row in df_pendentes.iterrows():
                    with st.container(border=True):
                        st.markdown(f"👤 **Cliente:** {row['usuario']}")
                        v_mensal = st.number_input("Definir Mensalidade (R$):", min_value=0.0, value=49.90, key=f"v_pend_{row['id']}")
                        if st.button("✅ Ativar Conta", key=f"btn_lib_{row['id']}", use_container_width=True, type="primary"):
                            if atualizar_status_e_mensalidade(row['id'], 'ativo', v_mensal):
                                st.success(f"Acesso liberado para {row['usuario']}!")
                                st.rerun()
        else:
            st.info("Nenhum registro encontrado.")
            
    with adm_tab2:
        st.write("### ➕ Cadastrar novo cliente")
        with st.form("form_cadastro_manual", clear_on_submit=True):
            novo_usr = st.text_input("Nome de Usuário:")
            nova_sen = st.text_input("Senha Inicial:", type="password")
            novo_tel = st.text_input("Telefone (com DDD):", placeholder="65999998888")
            mensalidade_manual = st.number_input("Valor da Mensalidade (R$):", min_value=0.0, value=49.90)
            status_manual = st.selectbox("Status Inicial:", ["ativo", "inativo", "pendente"])
            
            if st.form_submit_button("Salvar e Criar Cliente", width="stretch", type="primary"):
                if novo_usr and nova_sen:
                    res_manual = criar_usuario_rest(novo_usr, nova_sen, status=status_manual, valor_mensalidade=mensalidade_manual, telefone=novo_tel)
                    if res_manual == "Existe":
                        st.error("Este nome de usuário já existe.")
                    elif res_manual:
                        st.success(f"🎉 Cliente '{novo_usr}' cadastrado com sucesso!")
                        st.rerun()
                else:
                    st.error("Preencha o usuário e a senha.")

    with adm_tab3:
        st.write("### 👥 Gerenciamento de Clientes")
        if not df_users.empty:
            for idx, row in df_users.iterrows():
                if row.get('role') == 'admin':
                    continue
                
                status_atual = row.get('status', 'ativo')
                
                with st.container(border=True):
                    col_info1, col_info2 = st.columns([2, 1])
                    col_info1.markdown(f"👤 **{row['usuario']}**")
                    
                    if status_atual == 'ativo':
                        col_info2.markdown("🟢 **Ativo**")
                    elif status_atual == 'inativo':
                        col_info2.markdown("🔴 **Inativo**")
                    else:
                        col_info2.markdown("⏳ **Pendente**")
                    
                    st.caption(f"💵 Mensalidade: **{fmt_moeda(row.get('valor_mensalidade', 0.0))}**")
                    col_b1, col_b2 = st.columns(2)
                    
                    with col_b1:
                        if status_atual == 'ativo':
                            if st.button("🚫 Suspender", key=f"btn_susp_{row['id']}", width="stretch"):
                                atualizar_status_e_mensalidade(row['id'], 'inativo', float(row.get('valor_mensalidade', 0.0)))
                                st.rerun()
                        else:
                            if st.button("⚡ Reativar", key=f"btn_reat_{row['id']}", width="stretch"):
                                atualizar_status_e_mensalidade(row['id'], 'ativo', float(row.get('valor_mensalidade', 0.0)))
                                st.rerun()
                                
                    with col_b2:
                        if status_atual in ['inativo', 'pendente']:
                            tel_cadastro = row.get('telefone', '')
                            if tel_cadastro:
                                url_whatsapp = criar_link_cobranca(telefone=tel_cadastro, nome=row['usuario'], valor=float(row.get('valor_mensalidade', 0.0)))
                                st.link_button("💬 Cobrar", url_whatsapp, width="stretch")
                            else:
                                if st.button("❌ Excluir", key=f"btn_del_{row['id']}", width="stretch"):
                                    if excluir_usuario_admin(row['id']):
                                        st.success("Usuário deletado!")
                                        st.rerun()
                        else:
                            if st.button("❌ Excluir", key=f"btn_del_{row['id']}", width="stretch"):
                                if excluir_usuario_admin(row['id']):
                                    st.success("Usuário deletado!")
                                    st.rerun()

# --- ABA 1: DASHBOARD ---
elif opcao == "📊 Dashboard":
    st.markdown("<h2>📊 Dashboard Financeiro</h2>", unsafe_allow_html=True)
    
    opcoes_meses = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    
    col_mes, col_ano = st.columns(2)
    with col_mes:
        mes_nome = st.selectbox("Filtrar por Mês", opcoes_meses, index=datetime.now().month)
        mes_selecionado = "Todos" if mes_nome == "Todos" else f"{opcoes_meses.index(mes_nome):02d}"

    with col_ano:
        opcoes_anos = ["2026", "2027", "2028", "2029", "2030"]
        ano_atual_str = str(datetime.now().year)
        idx_ano = opcoes_anos.index(ano_atual_str) if ano_atual_str in opcoes_anos else 0
        ano_selecionado = st.selectbox("Filtrar por Ano", opcoes_anos, index=idx_ano)
    
    dados = dados_dashboard(st.session_state["usuario_id"], mes_selecionado, ano_selecionado)
    
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("💰 Total Receitas", fmt_moeda(dados['receitas']))
    m2.metric("💸 Total Despesas", fmt_moeda(dados['despesas']))
    m3.metric("🏦 Saldo Atual", fmt_moeda(dados['saldo']))
    m4.metric("🎯 Eficiência", f"{(dados['saldo']/dados['receitas']*100) if dados['receitas'] > 0 else 0:.1f}%")

    st.markdown("---")
    st.markdown("### 🤖 Assistente de IA / Insights Financeiros")

    try:
        metas_raw = listar_metas(st.session_state["usuario_id"])
        insights_gerados = []
        rec_tot = float(dados.get('receitas', 0))
        desp_tot = float(dados.get('despesas', 0))
        
        if rec_tot > 0:
            pct_comp = (desp_tot / rec_tot) * 100
            if pct_comp >= 85:
                insights_gerados.append({
                    "tipo": "error", "icone": "🚨", "titulo": "Alerta de Comprometimento",
                    "texto": f"Suas despesas representam **{pct_comp:.1f}%** da sua receita neste período!"
                })
            elif pct_comp <= 50:
                insights_gerados.append({
                    "tipo": "success", "icone": "👏", "titulo": "Excelente Gestão",
                    "texto": f"Você comprometeu apenas **{pct_comp:.1f}%** das suas receitas até agora."
                })

        if metas_raw:
            for m in metas_raw:
                try:
                    nome_m, alvo_m, guardado_m = m[1], float(m[2]), float(m[3])
                    if alvo_m > 0:
                        pct_m = (guardado_m / alvo_m) * 100
                        if pct_m >= 100:
                            insights_gerados.append({"tipo": "success", "icone": "🏆", "titulo": "Meta Alcançada!", "texto": f"Parabéns! A meta **{nome_m}** atingiu **100%** do objetivo!"})
                        elif pct_m >= 75:
                            insights_gerados.append({"tipo": "info", "icone": "🎯", "titulo": f"Meta {nome_m}", "texto": f"Falta pouco! Você já atingiu **{pct_m:.0f}%** da meta **{nome_m}**."})
                except Exception:
                    pass

        if insights_gerados:
            cols_ins = st.columns(min(len(insights_gerados), 3))
            for idx, item in enumerate(insights_gerados[:3]):
                with cols_ins[idx % len(cols_ins)]:
                    msg = f"{item['icone']} **{item['titulo']}**\n\n{item['texto']}"
                    if item["tipo"] == "warning": st.warning(msg)
                    elif item["tipo"] == "success": st.success(msg)
                    elif item["tipo"] == "error": st.error(msg)
                    else: st.info(msg)
        else:
            st.info("💡 **Tudo sob controle**: Não foram detectados alertas críticos para o filtro selecionado.")

    except Exception as e:
        st.error(f"⚠️ Erro ao processar os insights: {e}")

    st.markdown("### 🎯 Progresso do Orçamento Mensal")
    limite_definido = obter_limite_orcamento(st.session_state["usuario_id"])
    if limite_definido > 0:
        porcentagem_gasta = min(float(dados['despesas']) / float(limite_definido), 1.0)
        st.progress(porcentagem_gasta)
        restante = limite_definido - dados['despesas']
        if restante >= 0:
            st.success(f"Você utilizou **{porcentagem_gasta*100:.1f}%** do seu limite. Ainda restam **{fmt_moeda(restante)}** disponíveis.")
        else:
            st.error(f"⚠️ Atenção! Você **estourou** o seu orçamento em **{fmt_moeda(abs(restante))}**.")
    else:
        st.info("💡 Você ainda não definiu um teto de orçamento. Vá na aba '⚙️ Configurações' para estabelecer um limite.")

    st.markdown("---")
    c_graf1, c_graf2 = st.columns(2)
    with c_graf1:
        st.write("### 📈 Fluxo de Caixa Mensal")
        meses, recs, desps = dados_grafico_mensal(st.session_state["usuario_id"], ano_selecionado)
        fig_mensal = go.Figure()
        fig_mensal.add_trace(go.Bar(x=meses, y=recs, name='Receitas', marker_color='#2ecc71'))
        fig_mensal.add_trace(go.Bar(x=meses, y=desps, name='Despesas', marker_color='#e74c3c'))
        fig_mensal.update_layout(barmode='group', height=350, margin=dict(l=20, r=20, t=20, b=20))
        st.plotly_chart(fig_mensal, width="stretch")
    with c_graf2:
        st.write("### 🍕 Despesas por Categoria")
        cats, valores = dados_grafico_categorias(st.session_state["usuario_id"], mes_selecionado, ano_selecionado)
        if cats:
            fig_pizza = px.pie(names=cats, values=valores, hole=0.4)
            fig_pizza.update_layout(height=350, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig_pizza, width="stretch")
        else:
            st.info("Nenhuma despesa registrada para o período selecionado.")

    st.markdown("---")
    st.markdown("### 🏷️ Gastos por Tag / Evento")

    try:
        tags_retornadas, valores_tags = dados_grafico_tags(st.session_state["usuario_id"], mes_selecionado, ano_selecionado)
        if tags_retornadas and len(tags_retornadas) > 0:
            df_grouped = (
                pd.DataFrame({"Tag": tags_retornadas, "Valor": valores_tags})
                .assign(Tag_Norm=lambda x: "#" + x["Tag"].astype(str).str.strip().str.lower().str.lstrip("#"))
                .groupby("Tag_Norm", as_index=False)["Valor"].sum()
            )

            col_t1, col_t2 = st.columns([1.5, 1])
            with col_t1:
                fig_tags = px.pie(df_grouped, names="Tag_Norm", values="Valor", hole=0.4, color_discrete_sequence=px.colors.qualitative.Pastel)
                fig_tags.update_traces(textposition='inside', textinfo='percent+label')
                fig_tags.update_layout(height=350, margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig_tags, width="stretch")
                
            with col_t2:
                st.write("#### 📋 Detalhamento")
                df_tags_detalhe = pd.DataFrame({"Tag": df_grouped["Tag_Norm"], "Total Gasto": df_grouped["Valor"].apply(fmt_moeda)})
                st.dataframe(df_tags_detalhe, width="stretch")
        else:
            st.info("💡 Nenhuma movimentação com tag registrada para o período selecionado.")
    except Exception:
        pass

    st.markdown("---")
    st.markdown("### 🎯 Progresso das Metas de Economia")

    metas = listar_metas(st.session_state["usuario_id"])

    if metas:
        nomes_metas = [m[1] for m in metas]
        valores_guardados = [float(m[3]) for m in metas]
        valores_alvo = [float(m[2]) for m in metas]

        col_g1, col_g2 = st.columns(2)

        with col_g1:
            st.write("#### 📊 Comparativo por Meta")
            fig_metas = go.Figure()
            fig_metas.add_trace(go.Bar(x=nomes_metas, y=valores_guardados, name='Guardado (R$)', marker_color='#2ecc71'))
            fig_metas.add_trace(go.Bar(x=nomes_metas, y=valores_alvo, name='Objetivo (R$)', marker_color='#3498db'))
            fig_metas.update_layout(barmode='group', height=350, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig_metas, width="stretch")

        with col_g2:
            st.write("#### 🎯 Conclusão Geral dos Objetivos")
            total_guardado, total_alvo = sum(valores_guardados), sum(valores_alvo)
            porcentagem_total = (total_guardado / total_alvo * 100) if total_alvo > 0 else 0

            fig_gauge = go.Figure(go.Indicator(
                mode="gauge+number",
                value=porcentagem_total,
                number={'suffix': "%"},
                gauge={
                    'axis': {'range': [0, 100]},
                    'bar': {'color': "#2ecc71"},
                    'steps': [{'range': [0, 50], 'color': "#34495e"}, {'range': [50, 85], 'color': "#2980b9"}],
                }
            ))
            fig_gauge.update_layout(height=350, margin=dict(l=20, r=20, t=20, b=20))
            st.plotly_chart(fig_gauge, width="stretch")

    else:
        st.info("Nenhuma meta cadastrada até o momento. Cadastre suas metas no menu lateral!")

# --- ABA 2: METAS ---
elif opcao == "🎯 Metas de Economia":
    st.title("🎯 Metas de Economia")
    aba1, aba2 = st.tabs(["📋 Suas Metas", "➕ Criar Nova Meta"])
    
    with aba2:
        with st.form("nova_meta_form", clear_on_submit=True):
            nome_m = st.text_input("Objetivo")
            alvo_m = st.number_input("Valor Alvo (R$)", min_value=1.0)
            prazo_m = st.date_input("Prazo Limite", format="DD/MM/YYYY")
            
            if st.form_submit_button("Salvar Meta") and nome_m:
                criar_meta(st.session_state["usuario_id"], nome_m, alvo_m, prazo_m)
                st.cache_data.clear()
                st.success("Meta criada com sucesso!")
                st.rerun()
                
    with aba1:
        metas = listar_metas(st.session_state["usuario_id"])
        if not metas: 
            st.info("Você ainda não criou nenhuma meta de economia.")
        else:
            for m in metas:
                data_raw = str(m[4])
                try:
                    data_limite_fmt = datetime.strptime(data_raw.replace("/", "-").split("T")[0], "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    data_limite_fmt = data_raw

                st.write(f"### {m[1]} (Até: {data_limite_fmt})")
                
                def limpar_valor_numerico(val):
                    if val is None:
                        return 0.0
                    texto_limpo = re.sub(r"[^\d,. ]", "", str(val)).strip()
                    try:
                        if "," in texto_limpo:
                            texto_limpo = texto_limpo.replace(".", "").replace(",", ".")
                        return float(texto_limpo)
                    except Exception:
                        return 0.0

                val_guardado = limpar_valor_numerico(m[3])
                val_alvo = limpar_valor_numerico(m[2])

                guardado_str = f"R$ {val_guardado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                alvo_str = f"R$ {val_alvo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                st.markdown(
                    f"<p style='font-size: 1.1rem; margin-bottom: 0px;'><b>Guardado:</b> {guardado_str} de {alvo_str}</p>", 
                    unsafe_allow_html=True
                )
                
                progresso = min(val_guardado / val_alvo, 1.0) if val_alvo > 0 else 0.0
                st.progress(progresso)
                
                col_v, col_b1, col_b2, col_b3 = st.columns([2, 1, 1, 1])

                with col_v:
                    valor_mov = st.number_input("Valor da movimentação (R$):", min_value=0.0, value=100.0, step=50.0, format="%.2f", key=f"val_{m[0]}")

                with col_b1:
                    st.write("")
                    st.write("")
                    if st.button("➕ Depositar", key=f"dep_{m[0]}", width="stretch"):
                        novo_saldo = val_guardado + valor_mov
                        atualizar_progresso_meta(m[0], novo_saldo)
                        st.cache_data.clear()
                        mov_fmt = f"R$ {valor_mov:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        st.success(f"Guardado +{mov_fmt}!")
                        st.rerun()

                with col_b2:
                    st.write("")
                    st.write("")
                    if st.button("➖ Resgatar", key=f"res_{m[0]}", width="stretch"):
                        if valor_mov > val_guardado:
                            st.error("O valor de resgate é maior do que o saldo guardado!")
                        else:
                            novo_saldo = val_guardado - valor_mov
                            atualizar_progresso_meta(m[0], novo_saldo)
                            st.cache_data.clear()
                            mov_fmt = f"R$ {valor_mov:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                            st.success(f"Retirado -{mov_fmt}!")
                            st.rerun()

                with col_b3:
                    st.write("")
                    st.write("")
                    if st.button("🗑️ Excluir", key=f"del_{m[0]}", width="stretch", type="secondary"):
                        if excluir_meta(st.session_state["usuario_id"], m[0]):
                            st.cache_data.clear()
                            st.success(f"Meta '{m[1]}' excluída com sucesso!")
                            st.rerun()
                        else:
                            st.error("Erro ao excluir a meta no banco de dados.")
                
                st.markdown("---")
                
# --- ABA 3: CONTAS ---
elif opcao == "🏦 Gerir Contas":
    st.title("🏦 Gerenciamento de Contas Bancárias")
    
    OPCOES_BANCOS = [
        "260 - Nubank", "001 - Banco do Brasil", "341 - Itaú Unibanco", 
        "237 - Bradesco", "033 - Santander", "104 - Caixa Econômica Federal", 
        "336 - C6 Bank", "077 - Banco Inter", "380 - PicPay", 
        "000 - Carteira (Dinheiro)", "Outro"
    ]
    
    with st.form("cadastro_conta_form", clear_on_submit=True):
        banco_selecionado = st.selectbox("Escolha o Banco/Conta que deseja criar", OPCOES_BANCOS)
        nome_personalizado = st.text_input("Nome personalizado (Caso escolha 'Outro'):")
        saldo_inicial_input = st.number_input("Saldo Inicial (R$)", value=0.0, step=10.0)
        
        if st.form_submit_button("Cadastrar Conta"):
            nome_final = nome_personalizado.strip() if banco_selecionado == "Outro" else banco_selecionado
            if nome_final:
                if cadastrar_conta(st.session_state["usuario_id"], nome_final, saldo_inicial_input):
                    st.success(f"Conta '{nome_final}' cadastrada com sucesso!")
                    st.rerun()
                else:
                    st.error("Erro ao cadastrar conta. Tente novamente.")
            else:
                st.error("Por favor, informe o nome da conta para continuar.")

    st.markdown("---")
    st.subheader("📋 Contas Cadastradas")
    
    contas_lista = listar_contas(st.session_state["usuario_id"])
    
    if contas_lista:
        for item in contas_lista:
            if isinstance(item, dict):
                cid = item.get("id")
                cnome = item.get("nome", "Sem Nome")
                csaldo = float(item.get("saldo_inicial", 0.0))
            else:
                cid, cnome, csaldo = item[0], item[1], float(item[2])

            col_c1, col_c2, col_c3 = st.columns([3, 2, 1])
            col_c1.write(f"🔹 **{cnome}**")
            col_c2.write(f"Saldo Inicial: {formatar_moeda_ptbr(csaldo)}")
            
            with col_c3:
                if st.button("Remover", key=f"del_c_{cid}", width="stretch"):
                    if excluir_conta(st.session_state["usuario_id"], cid):
                        st.success("Conta removida com sucesso!")
                        st.rerun()
                    else:
                        st.error("Erro ao remover a conta selecionada.")
    else:
        st.info("Nenhuma conta cadastrada até o momento.")

# --- ABA 4: MOVIMENTAÇÕES ---
elif opcao == "💸 Lançar Movimentações":
    st.title("💸 Lançar Movimentações")
    
    contas_raw = listar_contas(st.session_state["usuario_id"]) or []
    cartoes_raw = listar_cartoes(st.session_state["usuario_id"]) or []

    mapa_contas = {}
    for c in contas_raw:
        if isinstance(c, dict):
            mapa_contas[c.get('nome', 'Conta')] = c['id']
        elif isinstance(c, (list, tuple)):
            mapa_contas[str(c[1])] = c[0]

    mapa_cartoes, cartoes_info = {}, {}
    for c in cartoes_raw:
        if isinstance(c, dict):
            label_cartao = f"{c['nome_cartao']} (Fecha dia {c['dia_fechamento']})"
            mapa_cartoes[label_cartao] = c['id']
            cartoes_info[c['id']] = c['dia_fechamento']

    modalidade = st.radio("Frequência do Lançamento", ["Único", "Parcelado", "Fixo / Recorrente"], horizontal=True)
    tipo_mov = st.selectbox("Tipo de Lançamento", ["Despesa", "Receita"])
    
    OPCOES_DESCRICAO = [
        "Mercado", "Combustível", "Custo de Casa", "Pix Recebido", "Pix Enviado",
        "Cartão de Débito", "Cartão de Crédito", "Almoço / Jantar / Lanche",
        "Pagamento de Salário", "Conta de Luz", "Assinatura (Netflix, Spotify, etc.)",
        "Conta de Água", "Internet", "Transferência entre Contas", "Outros"
    ]
    
    desc_container = st.container()
    with desc_container:
        desc_selecionada = st.selectbox("Descrição / Histórico", OPCOES_DESCRICAO, key="select_desc_mov")
        desc_customizada = ""
        if desc_selecionada == "Outros":
            desc_customizada = st.text_input("Digite a descrição personalizada:", placeholder="Ex: Compra na feira", key="txt_desc_custom")

    forma = st.selectbox("Forma de Pagamento", ["Pix", "Dinheiro", "Cartão de Crédito", "Cartão de Débito", "Boleto"])

    with st.form("lancamento_form", clear_on_submit=True, border=False):
        cartao_id_sel, fechamento_cartao_sel, conta_label = None, None, None

        if forma == "Cartão de Crédito":
            if mapa_cartoes:
                cartao_label = st.selectbox("Selecione o Cartão de Crédito", list(mapa_cartoes.keys()))
                cartao_id_sel = mapa_cartoes[cartao_label]
                fechamento_cartao_sel = cartoes_info.get(cartao_id_sel)
            else:
                st.warning("⚠️ Nenhum cartão encontrado! Por favor, cadastre um cartão na aba '💳 Cartões & Faturas' primeiro.")
        else:
            if mapa_contas:
                conta_label = st.selectbox("Conta Origem/Destino", list(mapa_contas.keys()))
            else:
                st.warning("⚠️ Nenhuma conta encontrada! Por favor, cadastre uma conta primeiro para lançar via " + forma + ".")

        col_v, col_q = st.columns(2)
        with col_v:
            val = st.number_input("Valor da Parcela (R$)" if modalidade == "Parcelado" else "Valor (R$)", min_value=0.0, step=50.0, format="%.2f")
            
        num_repeticoes = 1
        with col_q:
            if modalidade == "Parcelado":
                num_repeticoes = st.number_input("Quantidade de Parcelas", min_value=2, max_value=72, value=2, step=1)
            elif modalidade == "Fixo / Recorrente":
                num_repeticoes = st.number_input("Repetir por quantos meses?", min_value=2, max_value=60, value=12, step=1)

        cat_sel = st.selectbox("Categoria", CATEGORIAS_DESPADREVAL if tipo_mov == "Despesa" else CATEGORIAS_RECEITAS)
        tags_input = st.text_input("🏷️ Tags / Etiquetas (Opcional)", placeholder="Ex: #Viagem2026, #Trabalho, #Reforma", help="Separe por vírgulas para indicar eventos ou projetos específicos.")

        fuso_br = pytz.timezone("America/Sao_Paulo")
        hoje_br = datetime.now(fuso_br).date()
        data_f = st.date_input("Data da Operação", value=hoje_br, format="DD/MM/YYYY")
        
        enviado = st.form_submit_button("Registrar Transação")
        
        if enviado:
            desc_final = desc_customizada.strip() if desc_selecionada == "Outros" else desc_selecionada

            if desc_selecionada == "Outros" and not desc_final:
                st.error("Por favor, digite a descrição personalizada.")
            elif forma == "Cartão de Crédito" and not cartao_id_sel:
                st.error("Por favor, selecione ou cadastre um Cartão de Crédito antes de continuar.")
            elif forma != "Cartão de Crédito" and not conta_label:
                st.error("Por favor, selecione uma Conta Origem/Destino.")
            else:
                data_salvar = data_f.strftime("%Y-%m-%d")
                id_real_conta = mapa_contas[conta_label] if conta_label else None
                mes_fatura_calc = None
                if cartao_id_sel and fechamento_cartao_sel:
                    mes_fatura_calc = calcular_mes_fatura(data_salvar, fechamento_cartao_sel)

                if modalidade == "Parcelado":
                    sucesso = salvar_movimentacao_parcelada(
                        usuario_id=st.session_state["usuario_id"], conta_id=id_real_conta, descricao=desc_final,
                        valor=val, tipo=tipo_mov, forma_pagamento=forma, parcelas=int(num_repeticoes),
                        data_base=data_salvar, categoria=cat_sel, cartao_id=cartao_id_sel,
                        dia_fechamento=fechamento_cartao_sel, tags=tags_input
                    )
                elif modalidade == "Fixo / Recorrente":
                    sucesso = salvar_movimentacao_recorrente(
                        usuario_id=st.session_state["usuario_id"], conta_id=id_real_conta, descricao=desc_final,
                        valor=val, tipo=tipo_mov, forma_pagamento=forma, meses=int(num_repeticoes),
                        data_base=data_salvar, categoria=cat_sel, cartao_id=cartao_id_sel,
                        dia_fechamento=fechamento_cartao_sel, tags=tags_input
                    )
                else:
                    sucesso = salvar_movimentacao(
                        usuario_id=st.session_state["usuario_id"], conta_id=id_real_conta, descricao=desc_final,
                        valor=val, tipo=tipo_mov, forma_pagamento=forma, data_str=data_salvar,
                        categoria=cat_sel, cartao_id=cartao_id_sel, mes_fatura=mes_fatura_calc, tags=tags_input
                    )
                
                if sucesso:
                    st.cache_data.clear()
                    st.success(f"Transação ({modalidade}) salva com sucesso!")
                    st.rerun()
                else:
                    st.error("Erro ao salvar a transação no banco de dados. Tente novamente.")

    st.markdown("---")
    st.subheader("🕒 Lançamentos Recentes")
    
    dados_recentes = buscar_todas_movimentacoes(st.session_state["usuario_id"], "Todos", "Todos")
    
    if dados_recentes:
        df_recentes = pd.DataFrame(dados_recentes, columns=["ID", "Data", "Conta", "Tipo", "Forma Pagto", "Descrição", "Valor", "Categoria"])
        df_recentes["ID"] = pd.to_numeric(df_recentes["ID"])
        df_recentes = df_recentes.sort_values(by="ID", ascending=False).head(20)
        
        df_exibicao = df_recentes.copy()
        df_exibicao["Valor"] = df_exibicao["Valor"].apply(lambda v: fmt_moeda(v))
        df_exibicao["Data"] = pd.to_datetime(df_exibicao["Data"]).dt.strftime('%d/%m/%Y')
        
        with st.container(height=300):
            st.dataframe(df_exibicao, width="stretch", hide_index=True)

        st.markdown("#### 🗑️ Excluir Lançamento Incorreto")
        col_del_1, col_del_2 = st.columns([3, 1])
        
        with col_del_1:
            ids_disponiveis = [int(i) for i in df_recentes["ID"].tolist()]
            id_para_deletar = st.selectbox("Selecione o ID da transação que deseja remover:", ids_disponiveis, key="select_del_id")
            
        with col_del_2:
            st.write("")
            st.write("") 
            if st.button("🗑️ Excluir", width="stretch", type="secondary"):
                id_limpo = int(id_para_deletar)
                if excluir_movimentacao(usuario_id=st.session_state["usuario_id"], mov_id=id_limpo):
                    st.cache_data.clear()
                    st.success(f"Lançamento ID {id_limpo} excluído com sucesso!")
                    st.rerun()
                else:
                    st.error("Não foi possível excluir o lançamento selecionado.")
    else:
        st.info("Nenhuma movimentação cadastrada até o momento.")

# --- ABA 5: EXTRATO DETALHADO ---
elif opcao == "📋 Extrato Detalhado":
    st.title("📋 Extrato Detalhado de Transações")

    # Mês e Ano atuais automáticos
    hoje = datetime.now()
    mes_atual_idx = hoje.month  # 1 para Jan, 2 para Fev, ..., 8 para Ago, etc.
    ano_atual_str = str(hoje.year)

    opcoes_meses = [
        "Todos",
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]

    opcoes_anos = ["Todos", "2026", "2027", "2028", "2029", "2030"]
    idx_ano_atual = (
        opcoes_anos.index(ano_atual_str) if ano_atual_str in opcoes_anos else 0
    )

    col_e1, col_e2 = st.columns(2)
    with col_e1:
        mes_nome_extrato = st.selectbox(
            "Filtrar Mês", opcoes_meses, index=mes_atual_idx
        )  # Padrão: Mês Atual

    with col_e2:
        ano_extrato = st.selectbox(
            "Filtrar Ano",
            opcoes_anos,
            index=idx_ano_atual,
        )  # Padrão: Ano Atual

    # Reseta a página ao mudar filtros
    chave_filtro = f"{mes_nome_extrato}_{ano_extrato}"
    if st.session_state.get("ultimo_filtro") != chave_filtro:
        st.session_state["pagina_extrato"] = 1
        st.session_state["ultimo_filtro"] = chave_filtro

    if "pagina_extrato" not in st.session_state:
        st.session_state["pagina_extrato"] = 1

    ITENS_POR_PAGINA = 100

    dados_banco, total_registros, total_paginas = (
        buscar_movimentacoes_paginadas(
            usuario_id=st.session_state["usuario_id"],
            mes=mes_nome_extrato,
            ano=ano_extrato,
            pagina=st.session_state["pagina_extrato"],
            itens_por_pagina=ITENS_POR_PAGINA,
        )
    )

    if dados_banco:
        # CONVERSÃO: Transforma a lista de dicionários em tuplas para o Excel/PDF
        dados_tuplas = []
        for item in dados_banco:
            # Extrai o nome da conta se vier do relacionamento
            conta_nome = "Conta"
            if isinstance(item.get("contas"), dict):
                conta_nome = item["contas"].get("nome", "Conta")

            dados_tuplas.append(
                (
                    item.get("id"),
                    item.get("data"),
                    conta_nome,
                    item.get("tipo", ""),
                    item.get("forma_pagamento", ""),
                    item.get("descricao", ""),
                    item.get("valor", 0.0),
                    item.get("categoria", ""),
                )
            )

        st.write("### 📥 Exportar Relatórios")
        c_btn1, c_btn2, _ = st.columns([1.5, 1.5, 4])

        with c_btn1:
            # Passamos 'dados_tuplas' em vez de 'dados_banco'
            excel_data = gerar_excel_profissional(
                dados_tuplas, mes_nome_extrato, ano_extrato
            )
            st.download_button(
                "🟢 Baixar Excel (.xlsx)",
                data=excel_data,
                file_name=f"extrato_financeiro_{mes_nome_extrato}_{ano_extrato}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with c_btn2:
            # Passamos 'dados_tuplas' em vez de 'dados_banco'
            pdf_data = gerar_pdf_profissional(
                dados_tuplas, mes_nome_extrato, ano_extrato
            )
            st.download_button(
                "🔴 Baixar PDF Relatório",
                data=pdf_data,
                file_name=f"extrato_financeiro_{mes_nome_extrato}_{ano_extrato}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        st.markdown("---")

        # Exibição na tabela do Streamlit
        df = pd.DataFrame(dados_banco)

        if "contas" in df.columns:
            df["Conta"] = df["contas"].apply(
                lambda c: c.get("nome") if isinstance(c, dict) else "Conta"
            )
        else:
            df["Conta"] = "Conta"

        if "tags" not in df.columns:
            df["tags"] = ""

        todas_tags = set()
        for t in df["tags"].dropna():
            if t:
                todas_tags.update(
                    [x.strip() for x in str(t).split(",") if x.strip()]
                )

        if todas_tags:
            tags_sel = st.multiselect(
                "🏷️ Filtrar por Tag / Evento", sorted(list(todas_tags))
            )
            if tags_sel:
                df = df[
                    df["tags"]
                    .astype(str)
                    .apply(
                        lambda x: any(
                            tag.lower() in x.lower() for tag in tags_sel
                        )
                    )
                ]

        df_exibir = pd.DataFrame()
        df_exibir["ID"] = df["id"]
        df_exibir["Data"] = pd.to_datetime(df["data"]).dt.strftime("%d/%m/%Y")
        df_exibir["Conta"] = df["Conta"]
        df_exibir["Tipo"] = df.get("tipo", "")
        df_exibir["Forma Pagto"] = df.get("forma_pagamento", "")
        df_exibir["Descrição"] = df.get("descricao", "")
        df_exibir["Valor"] = df["valor"].apply(lambda v: fmt_moeda(v))
        df_exibir["Categoria"] = df.get("categoria", "")

        st.caption(
            f"Exibindo {len(df_exibir)} de {total_registros} registros salvos."
        )
        st.dataframe(df_exibir, use_container_width=True)

        # Controles de Paginação
        st.markdown("---")
        c_pag1, c_pag2, c_pag3 = st.columns([1, 2, 1])

        with c_pag1:
            if st.button(
                "⬅️ Anterior",
                disabled=(st.session_state["pagina_extrato"] <= 1),
            ):
                st.session_state["pagina_extrato"] -= 1
                st.rerun()

        with c_pag2:
            st.markdown(
                f"<h5 style='text-align: center;'>Página {st.session_state['pagina_extrato']} de {total_paginas}</h5>",
                unsafe_allow_html=True,
            )

        with c_pag3:
            if st.button(
                "Próximo ➡️",
                disabled=(st.session_state["pagina_extrato"] >= total_paginas),
            ):
                st.session_state["pagina_extrato"] += 1
                st.rerun()

        st.markdown("---")
        st.write("#### 🛠️ Operações Avançadas")
        id_excluir = st.number_input(
            "Deseja deletar algum lançamento? Digite o ID dele aqui:",
            min_value=0,
            step=1,
        )
        if st.button("Excluir Lançamento") and id_excluir > 0:
            if excluir_movimentacao(st.session_state["usuario_id"], id_excluir):
                st.success(
                    f"Movimentação {id_excluir} excluída com sucesso!"
                )
                st.rerun()
            else:
                st.error(
                    "Não foi possível excluir a movimentação informada."
                )
    else:
        st.info("Nenhuma movimentação encontrada para o período filtrado.")

# --- ABA 6: CONFIGURAÇÕES ---
elif opcao == "⚙️ Configurações":
    st.title("⚙️ Painel de Configurações")
    st.write("### 🎯 Orçamento Global de Despesas")
    limite_atual = obter_limite_orcamento(st.session_state["usuario_id"])
    st.write(f"Seu Limite de Gastos Alvo Atual: **{fmt_moeda(limite_atual)}**")
    
    n_limite = st.number_input("Definir novo teto de orçamento (R$):", value=float(limite_atual))
    if st.button("Atualizar Limite"):
        if definir_limite_orcamento(st.session_state["usuario_id"], n_limite):
            st.success("Teto do orçamento atualizado com sucesso!")
            st.rerun()
        else:
            st.error("Erro ao atualizar o teto do orçamento.")
        
    st.markdown("---")
    st.write("### 🔒 Alteração de Credenciais")
    with st.form("alterar_senha_form"):
        nova_senha = st.text_input("Nova Senha de Acesso:", type="password")
        if st.form_submit_button("Alterar Senha"):
            if nova_senha:
                senha_hash = hashlib.sha256(nova_senha.encode('utf-8')).hexdigest()
                if alterar_senha_usuario(st.session_state["usuario_id"], senha_hash):
                    st.success("Senha modificada com sucesso!")
                else:
                    st.error("Erro operacional ao atualizar senha.")
            else:
                st.error("Digite uma senha válida antes de salvar.")

# --- ABA: ORÇAMENTOS POR CATEGORIA ---
elif opcao == "🎯 Orçamentos por Categoria":
    st.title("📊 Dashboard de Orçamentos e Limites")

    hoje = datetime.now().date()
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        mes_sel = st.selectbox(
            "📅 Mês", 
            options=list(range(1, 13)), 
            index=hoje.month - 1, 
            format_func=lambda m: ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"][m-1]
        )
    with col_f2:
        ano_sel = st.number_input("📆 Ano", min_value=2020, max_value=2035, value=hoje.year)

    uid = st.session_state["usuario_id"]
    limite_geral = obter_limite_orcamento(uid)
    limites_dict = obter_limites_por_categoria(uid)
    movs_todas = buscar_todas_movimentacoes(uid, "Todos", "Todos")

    gastos_por_cat = {}
    if movs_todas:
        df_m = pd.DataFrame(movs_todas, columns=["ID", "Data", "Conta", "Tipo", "Forma Pagto", "Descrição", "Valor", "Categoria"])
        df_despesas = df_m[df_m["Tipo"] == "Despesa"].copy()
        if not df_despesas.empty:
            df_despesas["Data"] = pd.to_datetime(df_despesas["Data"], errors="coerce")
            df_despesas = df_despesas[(df_despesas["Data"].dt.month == mes_sel) & (df_despesas["Data"].dt.year == ano_sel)]
            df_despesas["Valor"] = pd.to_numeric(df_despesas["Valor"], errors="coerce")
            gastos_por_cat = df_despesas.groupby("Categoria")["Valor"].sum().to_dict()

    total_orcado = sum(v if isinstance(v, (int, float)) else v.get("limite", 0.0) for v in limites_dict.values())
    total_gasto = sum(gastos_por_cat.get(cat, 0.0) for cat in limites_dict.keys())
    saldo_restante = total_orcado - total_gasto

    col_m1, col_m2, col_m3 = st.columns(3)
    col_m1.metric("🎯 Total Orçado", fmt_moeda(total_orcado))
    col_m2.metric("💸 Total Gasto", fmt_moeda(total_gasto))
    col_m3.metric("💰 Saldo Restante", fmt_moeda(saldo_restante), delta=fmt_moeda(saldo_restante), delta_color="normal" if saldo_restante >= 0 else "inverse")

    if limite_geral > 0:
        st.info(f"💡 Seu teto global cadastrado no sistema é de **{fmt_moeda(limite_geral)}**")

    st.markdown("---")
    col_cad, col_vis = st.columns([1, 2])

    if limites_dict:
        st.subheader("📈 Visão Geral dos Orçamentos")
        dados_grafico = []
        for cat, dados in limites_dict.items():
            lim = float(dados.get("limite", 0.0)) if isinstance(dados, dict) else float(dados)
            gst = float(gastos_por_cat.get(cat, 0.0))
            dados_grafico.append({"Categoria": cat, "Gasto Atual": gst, "Limite": lim})

        df_chart = pd.DataFrame(dados_grafico)

        if not df_chart.empty:
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                st.markdown("**Comparativo: Gasto vs. Limite**")
                st.bar_chart(df_chart.set_index("Categoria")[["Gasto Atual", "Limite"]], height=250)
            with col_g2:
                st.markdown("**Distribuição do Teto Orçado**")
                st.bar_chart(df_chart.set_index("Categoria")["Limite"], horizontal=True, height=250)

        st.markdown("---")

    with col_cad:
        with st.form("form_orcamento", clear_on_submit=True):
            st.subheader("⚙️ Definir Limite")
            cat_orc = st.selectbox("Categoria", CATEGORIAS_DESPADREVAL)
            limite_val = st.number_input("Teto Mensal (R$)", min_value=10.0, value=500.0, step=50.0)

            if st.form_submit_button("Salvar Limite", width="stretch"):
                if salvar_orcamento_categoria(uid, cat_orc, limite_val):
                    st.cache_data.clear()
                    st.success(f"Limite para '{cat_orc}' atualizado com sucesso!")
                    st.rerun()
                else:
                    st.error("Erro ao salvar limite no banco de dados. Tente novamente.")

    with col_vis:
        st.subheader("📈 Acompanhamento de Gastos")
        if not limites_dict:
            st.info("Nenhum limite por categoria cadastrado ainda. Defina um no formulário ao lado!")
        else:
            for cat_nome, dados_limite in limites_dict.items():
                if isinstance(dados_limite, dict):
                    limite = float(dados_limite.get("limite", 0.0))
                else:
                    limite = float(dados_limite)

                gasto_atual = float(gastos_por_cat.get(cat_nome, 0.0))
                porcentagem = min(gasto_atual / limite, 1.0) if limite > 0 else 0.0

                gasto_num = f"{gasto_atual:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                limite_num = f"{limite:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                if gasto_atual > limite:
                    exc = f"{(gasto_atual - limite):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    status = f"🔴 **ESTOURADO!** Excedeu em R$ {exc}"
                elif porcentagem >= 0.85:
                    rest = f"{(limite - gasto_atual):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    status = f"🟡 **Atenção!** Restam apenas R$ {rest}"
                else:
                    rest = f"{(limite - gasto_atual):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    status = f"🟢 **Dentro do Limite.** Restam R$ {rest}"

                # --- MANTIDO VISÍVEL NA TELA ---
                st.markdown(f"### 📌 {cat_nome}")
                st.markdown(f"<p style='font-size: 1rem; margin-bottom: 4px;'><b>Gasto:</b> R$ {gasto_num} de R$ {limite_num}</p>", unsafe_allow_html=True)
                st.progress(porcentagem)
                st.caption(status)

                # --- SETA EXPANSÍVEL PARA ALTERAR/EXCLUIR ---
                with st.expander("⚙️ Editar ou Excluir Orçamento"):
                    col_e1, col_e2 = st.columns(2)
                    novo_teto = col_e1.number_input("Alterar Limite (R$)", min_value=10.0, value=limite, step=50.0, key=f"edit_{cat_nome}")

                    if col_e1.button("💾 Salvar Alteração", key=f"btn_save_{cat_nome}"):
                        if salvar_orcamento_categoria(uid, cat_nome, novo_teto):
                            st.cache_data.clear()
                            st.success(f"Limite de {cat_nome} atualizado com sucesso!")
                            st.rerun()
                        else:
                            st.error("Erro ao alterar o limite selecionado.")

                    if col_e2.button("🗑️ Excluir Orçamento", key=f"btn_del_{cat_nome}", type="secondary"):
                        if excluir_orcamento_categoria(uid, cat_nome):
                            st.cache_data.clear()
                            st.success(f"Orçamento de {cat_nome} removido com sucesso!")
                            st.rerun()
                        else:
                            st.error("Erro ao excluir o orçamento selecionado.")

                st.markdown("---")

            
# --- ABA: PRÓXIMOS VENCIMENTOS ---
elif opcao == "📅 Próximos Vencimentos":
    st.title("📅 Próximos Vencimentos e Faturas")
    st.caption("Acompanhe contas pendentes e faturas de cartão agrupadas pelo ciclo de fechamento.")

    import calendar

    hoje = pd.to_datetime("today")
    mes_atual, ano_atual = hoje.month, hoje.year

    opcao_periodo = st.selectbox(
        "Selecione o Mês de Visualização:",
        options=[
            f"📅 Mês Atual ({mes_atual:02d}/{ano_atual})",
            f"⏪ Mês Anterior ({(mes_atual-1 if mes_atual > 1 else 12):02d}/{(ano_atual if mes_atual > 1 else ano_atual-1)})",
            f"⏩ Próximo Mês ({(mes_atual+1 if mes_atual < 12 else 1):02d}/{(ano_atual if mes_atual < 12 else ano_atual+1)})",
            "🔮 Ver Todos os Registros Encontrados",
        ],
    )

    if "Mês Atual" in opcao_periodo: target_mes, target_ano = mes_atual, ano_atual
    elif "Mês Anterior" in opcao_periodo: target_mes, target_ano = (mes_atual - 1 if mes_atual > 1 else 12), (ano_atual if mes_atual > 1 else ano_atual - 1)
    elif "Próximo Mês" in opcao_periodo: target_mes, target_ano = (mes_atual + 1 if mes_atual < 12 else 1), (ano_atual if mes_atual < 12 else ano_atual + 1)
    else: target_mes, target_ano = None, None

    usuario_atual_id = st.session_state.get("usuario_id")

    def normalizar_id(val):
        if pd.isna(val) or val is None: return ""
        s = str(val).strip()
        if s.lower() in ["none", "nan", "null", ""]: return ""
        try: return str(int(float(s)))
        except ValueError: return s.split(".")[0]

    def normalizar_mes_fatura(val):
        if pd.isna(val) or not val: return None, None
        if hasattr(val, 'month') and hasattr(val, 'year'):
            return int(val.month), int(val.year)
        s = str(val).strip()
        if "/" in s:
            partes = s.split("/")
            try: return int(partes[0]), int(partes[1])
            except Exception: return None, None
        elif "-" in s:
            partes = s.split("-")
            try:
                return (int(partes[1]), int(partes[0])) if len(partes[0]) == 4 else (int(partes[0]), int(partes[1]))
            except Exception: return None, None
        return None, None

    vencimentos = buscar_vencimentos_proximos(usuario_atual_id, dias=60)

    if not vencimentos:
        st.info("Nenum lançamento encontrado no banco de dados.")
    else:
        df_raw = pd.DataFrame(vencimentos)

        if "usuario_id" in df_raw.columns and usuario_atual_id is not None:
            usr_target = normalizar_id(usuario_atual_id)
            df_raw["_usr_norm"] = df_raw["usuario_id"].apply(normalizar_id)
            if usr_target:
                df_raw = df_raw[df_raw["_usr_norm"] == usr_target].copy()

        if df_raw.empty:
            st.warning("Nenhum lançamento pertence ao usuário logado.")
        else:
            df_raw["pago"] = df_raw["pago"].fillna(False).astype(bool) if "pago" in df_raw.columns else False
            if "cartao_id" not in df_raw.columns: df_raw["cartao_id"] = None
            if "forma_pagamento" not in df_raw.columns: df_raw["forma_pagamento"] = "N/A"

            mapa_cartoes_info = {}
            try:
                res_cartoes = listar_cartoes(usuario_atual_id)
                if res_cartoes:
                    for c in res_cartoes:
                        if isinstance(c, dict):
                            c_id = normalizar_id(c.get("id"))
                            mapa_cartoes_info[c_id] = {
                                "nome": str(c.get("nome_cartao") or c.get("nome") or "Cartão"),
                                "vencimento": int(c.get("dia_vencimento") or c.get("vencimento") or 15),
                                "fechamento": int(c.get("dia_fechamento") or c.get("fechamento") or 8),
                            }
            except Exception:
                pass

            # 1. IDENTIFICA APENAS LANÇAMENTOS DE CARTÃO DE CRÉDITO
            def eh_cartao(row):
                cid = normalizar_id(row.get("cartao_id"))
                fp = str(row.get("forma_pagamento") or "").lower()
                nc = str(row.get("nome_cartao") or "").strip()
                desc = str(row.get("descricao") or "").lower()
                cat = str(row.get("categoria") or "").lower()

                # Exclusão explícita para evitar que Pix/Boleto entrem como cartão
                if any(x in fp for x in ["pix", "boleto", "debito", "débito", "dinheiro", "especie", "espécie", "transferencia", "transferência"]):
                    return False

                if cid != "": return True
                if nc != "" and nc.lower() not in ["none", "nan"]: return True
                if any(x in fp for x in ["cart", "credito", "crédito", "fatura"]): return True
                if any(x in desc for x in ["cartão", "cartao", "credito", "crédito", "fatura"]): return True
                if any(x in cat for x in ["cartão", "cartao", "crédito"]): return True

                return False

            cond_cartao = df_raw.apply(eh_cartao, axis=1)

            # 2. FILTRA "OUTRAS CONTAS": MANTÉM SOMENTE BOLETOS E CONTAS FIXAS/RECORRENTES
            def eh_boleto_ou_recorrente(row):
                fp = str(row.get("forma_pagamento") or "").lower()
                desc = str(row.get("descricao") or "").lower()
                cat = str(row.get("categoria") or "").lower()
                tipo = str(row.get("tipo") or "").lower()

                eh_boleto = "boleto" in fp or "boleto" in desc or "boleto" in cat

                flag_recorrente = row.get("recorrente") or row.get("fixo") or row.get("e_fixo") or row.get("is_recorrente")
                eh_fixo_recorrente = (
                    bool(flag_recorrente) is True
                    or any(x in fp for x in ["fixo", "recorrente"])
                    or any(x in desc for x in ["fixo", "recorrente", "assinatura", "mensalidade"])
                    or any(x in cat for x in ["fixo", "recorrente", "assinatura", "mensalidade"])
                    or any(x in tipo for x in ["fixo", "recorrente"])
                )

                return eh_boleto or eh_fixo_recorrente

            df_outras_contas = df_raw[~cond_cartao].copy()
            if not df_outras_contas.empty:
                cond_permitidos = df_outras_contas.apply(eh_boleto_ou_recorrente, axis=1)
                df_outras_contas = df_outras_contas[cond_permitidos].copy()

                if not df_outras_contas.empty:
                    df_outras_contas["ids_compras"] = df_outras_contas["id"].apply(lambda x: [x])
                    df_outras_contas["descricao_detalhada"] = df_outras_contas["descricao"]

            # 3. TRATAMENTO DAS FATURAS DE CARTÃO
            df_credito = df_raw[cond_cartao].copy()
            df_faturas_agrupadas = pd.DataFrame()

            if not df_credito.empty:
                def obter_nome_cartao(row):
                    cid = normalizar_id(row.get("cartao_id"))
                    if cid in mapa_cartoes_info: return mapa_cartoes_info[cid]["nome"]
                    if "nome_cartao" in row and pd.notna(row["nome_cartao"]): return str(row["nome_cartao"])
                    return "Cartão"

                def resolver_datas_fatura(row):
                    cid = normalizar_id(row.get("cartao_id"))
                    info = mapa_cartoes_info.get(cid, {"vencimento": 15, "fechamento": 8})
                    dia_venc, dia_fech = info["vencimento"], info["fechamento"]

                    m_fat, a_fat = normalizar_mes_fatura(row.get("mes_fatura"))
                    if m_fat and a_fat:
                        mes_fatura, ano_fatura = m_fat, a_fat
                    else:
                        dt_compra = pd.to_datetime(row.get("data") or pd.to_datetime("today"))
                        ano, mes, dia_compra = dt_compra.year, dt_compra.month, dt_compra.day

                        if dia_compra >= dia_fech:
                            mes_fatura = mes + 1 if mes < 12 else 1
                            ano_fatura = ano if mes < 12 else ano + 1
                        else:
                            mes_fatura, ano_fatura = mes, ano

                    mes_fat_str = f"{mes_fatura:02d}/{ano_fatura:04d}"
                    max_dias = calendar.monthrange(ano_fatura, mes_fatura)[1]
                    dia_valido = min(dia_venc, max_dias)
                    data_venc_str = f"{ano_fatura:04d}-{mes_fatura:02d}-{dia_valido:02d}"

                    return pd.Series([data_venc_str, mes_fat_str])

                df_credito["nome_exibicao_cartao"] = df_credito.apply(obter_nome_cartao, axis=1)
                df_credito[["data_venc_calculada", "mes_fatura_calculado"]] = df_credito.apply(resolver_datas_fatura, axis=1)

                df_credito["data"] = df_credito["data_venc_calculada"]
                df_credito["mes_fatura"] = df_credito["mes_fatura_calculado"]
                df_credito["ids_compras"] = df_credito["id"]
                df_credito["descricao"] = df_credito["descricao"].fillna("Compra no Cartão") if "descricao" in df_credito.columns else "Compra no Cartão"

                df_faturas_agrupadas = (
                    df_credito.groupby(["nome_exibicao_cartao", "mes_fatura", "pago"], as_index=False, dropna=False)
                    .agg({
                        "valor": "sum", "id": "first", "ids_compras": lambda x: list(x),
                        "descricao": lambda x: list(x), "data": "first",
                        "categoria": lambda x: "Fatura de Cartão", "forma_pagamento": lambda x: "Cartão de Crédito",
                    })
                )

                df_faturas_agrupadas["descricao_detalhada"] = df_faturas_agrupadas["descricao"]
                df_faturas_agrupadas["descricao"] = df_faturas_agrupadas.apply(lambda r: f"💳 Fatura {r['nome_exibicao_cartao']} ({r['mes_fatura']})", axis=1)

            dfs_concatenar = [df for df in [df_outras_contas, df_faturas_agrupadas] if not df.empty]
            df_venc = pd.concat(dfs_concatenar, ignore_index=True) if dfs_concatenar else pd.DataFrame()

            if not df_venc.empty and target_mes and target_ano:
                def pertence_ao_periodo(row):
                    m_fat, a_fat = normalizar_mes_fatura(row.get("mes_fatura"))
                    if m_fat and a_fat: return m_fat == target_mes and a_fat == target_ano
                    if "data" in row and pd.notna(row["data"]):
                        dt = pd.to_datetime(row["data"])
                        return dt.month == target_mes and dt.year == target_ano
                    return False

                df_venc = df_venc[df_venc.apply(pertence_ao_periodo, axis=1)].copy()

            if df_venc.empty:
                st.info(f"Nenhum lançamento encontrado para o período {target_mes:02d}/{target_ano}.")
            else:
                df_venc = df_venc.sort_values(by="data").reset_index(drop=True)
                df_pendentes = df_venc[df_venc["pago"] == False].copy()
                df_pagos = df_venc[df_venc["pago"] == True].copy()

                col_kpi1, col_kpi2, col_kpi3 = st.columns(3)
                col_kpi1.metric("Total Pendente (A Pagar)", fmt_moeda(df_pendentes["valor"].sum()) if not df_pendentes.empty else "R$ 0,00")
                col_kpi2.metric("Total Já Pago (Quitado)", fmt_moeda(df_pagos["valor"].sum()) if not df_pagos.empty else "R$ 0,00")
                col_kpi3.metric("Total Geral do Mês", fmt_moeda(df_venc["valor"].sum()))

                st.markdown("---")
                tab_pendentes, tab_pagas = st.tabs([f"⏳ Contas Não Pagas / Pendentes ({len(df_pendentes)})", f"✅ Contas Pagas / Quitadas ({len(df_pagos)})"])

                with tab_pendentes:
                    if df_pendentes.empty:
                        st.success("🎉 Nenhuma conta pendente para este período!")
                    else:
                        for idx, row in df_pendentes.iterrows():
                            id_lanc = row["id"]
                            lista_ids = row.get("ids_compras") if isinstance(row.get("ids_compras"), list) else [id_lanc]
                            valor_fmt = fmt_moeda(row["valor"])
                            data_fmt = pd.to_datetime(row["data"]).strftime("%d/%m/%Y")

                            with st.expander(f"📅 {data_fmt} — {row['descricao']} | {valor_fmt}"):
                                st.write(f"**Categoria:** {row.get('categoria', 'Geral')}")
                                st.write(f"**Forma de Pagamento:** {row.get('forma_pagamento', 'N/A')}")

                                if isinstance(row.get("descricao_detalhada"), list):
                                    st.markdown("**🛒 Compras na fatura:**")
                                    for item in row["descricao_detalhada"]: st.caption(f"• {item}")

                                if st.button("✅ Marcar como Pago", key=f"pago_{idx}_{id_lanc}", type="primary"):
                                    for sub_id in lista_ids: marcar_lancamento_como_pago(sub_id)
                                    st.cache_data.clear()
                                    st.rerun()

                with tab_pagas:
                    if df_pagos.empty:
                        st.info("Nenhuma conta marcada como paga neste período.")
                    else:
                        for idx, row in df_pagos.iterrows():
                            id_lanc = row["id"]
                            lista_ids = row.get("ids_compras") if isinstance(row.get("ids_compras"), list) else [id_lanc]
                            valor_fmt = fmt_moeda(row["valor"])
                            data_fmt = pd.to_datetime(row["data"]).strftime("%d/%m/%Y")

                            with st.expander(f"✔️ {data_fmt} — {row['descricao']} | {valor_fmt}"):
                                st.write("**Status:** ✅ Quitado")

                                if isinstance(row.get("descricao_detalhada"), list):
                                    st.markdown("**🛒 Compras na fatura:**")
                                    for item in row["descricao_detalhada"]: st.caption(f"• {item}")

                                if st.button("↩️ Desfazer Pagamento", key=f"undo_{idx}_{id_lanc}"):
                                    for sub_id in lista_ids: desfazer_pagamento_lancamento(sub_id)
                                    st.cache_data.clear()
                                    st.rerun()

# --- ABA: CARTÕES & FATURAS ---
elif opcao == "💳 Cartões & Faturas":
    st.title("💳 Gestão de Cartões de Crédito & Faturas")

    tab_faturas, tab_novo_cartao, tab_gerenciar = st.tabs(["📄 Minhas Faturas", "➕ Cadastrar Novo Cartão", "⚙️ Gerenciar Cartões"])

    user_id = st.session_state.get("usuario_id")
    cartoes = listar_cartoes(user_id)

    with tab_faturas:
        if cartoes:
            c1, c2, c3 = st.columns(3)

            with c1:
                dict_cartoes = {c["id"]: c.get("nome_cartao") or c.get("nome") for c in cartoes}
                cartao_id_sel = st.selectbox("Escolha o Cartão", options=list(dict_cartoes.keys()), format_func=lambda x: dict_cartoes[x], key="sel_cartao_fatura")
                cartao_info = next(c for c in cartoes if c["id"] == cartao_id_sel)

            with c2:
                meses = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
                mes_fatura = st.selectbox("Mês da Fatura", meses, index=datetime.now().month - 1)

            with c3:
                ano_fatura = st.selectbox("Ano da Fatura", ["2026", "2027", "2028"], index=0)

            fatura_ref = f"{mes_fatura}/{ano_fatura}"
            st.markdown("---")

            dia_fechamento = int(cartao_info['dia_fechamento'])
            mes_int, ano_int = int(mes_fatura), int(ano_fatura)

            data_fim_fatura = datetime(ano_int, mes_int, dia_fechamento) - timedelta(days=1)
            data_inicio_fatura = datetime(ano_int - 1, 12, dia_fechamento) if mes_int == 1 else datetime(ano_int, mes_int - 1, dia_fechamento)

            periodo_str = f"{data_inicio_fatura.strftime('%d/%m/%Y')} a {data_fim_fatura.strftime('%d/%m/%Y')}"

            compras = buscar_gastos_fatura(user_id, cartao_id_sel, fatura_ref)
            total_fatura = sum(float(item["valor"]) for item in compras) if compras else 0.0

            todas_compras = buscar_gastos_fatura(user_id, cartao_id_sel, None)
            total_devedor_geral = sum(float(item["valor"]) for item in todas_compras if not item.get("pago", False) and not item.get("paga", False)) if todas_compras else 0.0

            limite_total = float(cartao_info["limite"])
            limite_disponivel = limite_total - total_devedor_geral

            st.info(f"💡 **Informações:** Fechamento todo **dia {cartao_info['dia_fechamento']}** | Vencimento todo **dia {cartao_info['dia_vencimento']}**")

            m1, m2, m3 = st.columns(3)
            m1.metric("Total da Fatura", formatar_moeda_ptbr(total_fatura))
            m2.metric("Limite Disponível", formatar_moeda_ptbr(limite_disponivel))
            m3.metric("Limite Total", formatar_moeda_ptbr(limite_total))

            st.write(f"### 🛒 Compras da Fatura ({fatura_ref})")
            st.warning(f"📆 **Atenção:** Esta fatura contempla as compras realizadas no período de **{periodo_str}**.\n\n💡 *Compras a partir do dia **{data_fim_fatura.day + 1:02d}/{mes_fatura}/{ano_fatura}** entram automaticamente na fatura do mês seguinte.*")
            
            if compras:
                fatura_paga = all(item.get("pago", False) or item.get("paga", False) for item in compras)

                if fatura_paga:
                    st.success(f"🎉 **Fatura Paga!** A fatura de {fatura_ref} (período {periodo_str}) já foi baixada e está quitada.")
                else:
                    if st.button("✅ Dar Baixa / Pagar Fatura Completa", type="primary"):
                        if dar_baixa_fatura_completa(user_id, cartao_id_sel, fatura_ref):
                            st.success(f"Fatura {fatura_ref} marcada como paga com sucesso!")
                            st.rerun()
                        else:
                            st.error("Erro ao dar baixa na fatura completa.")

                st.dataframe(compras, column_order=["data", "descricao", "categoria", "tags", "valor", "pago"], column_config={"valor": st.column_config.NumberColumn("Valor", format="%.2f")}, width="stretch")

                with st.expander("🗑️ Excluir um item desta fatura"):
                    dict_compras_excluir = {item["id"]: f"{item['data']} | {item['descricao']} - R$ {float(item['valor']):.2f}" for item in compras}
                    id_para_excluir = st.selectbox("Selecione o lançamento que deseja remover:", options=list(dict_compras_excluir.keys()), format_func=lambda x: dict_compras_excluir[x], key="select_excluir_fatura_item")
                    
                    if st.button("Confirmar Exclusão do Item", type="secondary"):
                        if excluir_movimentacao(user_id, id_para_excluir):
                            st.success("Lançamento excluído com sucesso!")
                            st.rerun()
                        else:
                            st.error("Erro ao excluir o lançamento no banco de dados.")

            else:
                st.warning(f"Nenhum gasto encontrado para a fatura de {fatura_ref}.")

        else:
            st.info("Nenhum cartão cadastrado. Use a aba ao lado para cadastrar seu primeiro cartão!")

    with tab_novo_cartao:
        st.subheader("➕ Adicionar Novo Cartão de Crédito")
        with st.form("form_cartao"):
            opcoes_bancos = ["260 - Nubank", "001 - Banco do Brasil", "341 - Itaú Unibanco", "237 - Bradesco", "033 - Santander", "104 - Caixa Econômica Federal", "336 - C6 Bank", "077 - Banco Inter", "380 - PicPay", "208 - BTG Pactual", "102 - XP Investimentos", "Outro"]
            nome_cartao_selecionado = st.selectbox("Nome do Cartão", options=opcoes_bancos, index=0)
            nome_outro = st.text_input("Se selecionou 'Outro', digite o nome do cartão:")
            nome_c = nome_outro if nome_cartao_selecionado == "Outro" else nome_cartao_selecionado
            limite_c = st.number_input("Limite de Crédito Total (R$)", min_value=0.0, value=1000.0, step=100.0)

            c_f, c_v = st.columns(2)
            with c_f: fechamento_c = st.number_input("Dia do Fechamento", min_value=1, max_value=31, value=20)
            with c_v: vencimento_c = st.number_input("Dia do Vencimento", min_value=1, max_value=31, value=30)

            btn_salvar = st.form_submit_button("Salvar Cartão")
            if btn_salvar:
                if nome_c.strip():
                    if cadastrar_cartao(user_id, nome_c, limite_c, fechamento_c, vencimento_c):
                        st.success(f"Cartão '{nome_c}' cadastrado com sucesso!")
                        st.rerun()
                    else: st.error("Erro ao cadastrar cartão no banco de dados.")
                else: st.error("Por favor, informe ou digite o nome do cartão.")

    with tab_gerenciar:
        st.subheader("⚙️ Alterar Limite ou Excluir Cartão")
        if cartoes:
            dict_cartoes_g = {c["id"]: c.get("nome_cartao") or c.get("nome") for c in cartoes}
            cartao_id_ger = st.selectbox("Selecione o Cartão para Configurar", options=list(dict_cartoes_g.keys()), format_func=lambda x: dict_cartoes_g[x], key="sel_cartao_gerenciar")
            cartao_sel_info = next(c for c in cartoes if c["id"] == cartao_id_ger)
            
            st.markdown("---")
            col_lim1, col_lim2 = st.columns([2, 1])
            with col_lim1:
                novo_limite = st.number_input("Novo Limite Total (R$)", min_value=0.0, value=float(cartao_sel_info["limite"]), step=100.0, key="input_novo_limite")
            with col_lim2:
                st.write("")
                st.write("")
                if st.button("✏️ Atualizar Limite", width="stretch"):
                    if atualizar_limite_cartao(user_id, cartao_id_ger, novo_limite):
                        st.success("Limite atualizado com sucesso!")
                        st.rerun()
                    else: st.error("Erro ao atualizar o limite.")

            st.markdown("---")
            st.warning("⚠️ **Zona de Perigo:** Excluir um cartão apaga as configurações dele.")
            if st.button("🗑️ Excluir Cartão", type="primary"):
                if excluir_cartao(user_id, cartao_id_ger):
                    st.success("Cartão excluído com sucesso!")
                    st.rerun()
                else: st.error("Erro ao excluir o cartão.")
        else:
            st.info("Nenhum cartão cadastrado para gerenciar no momento.")

# --- ABA: CONTAS A RECEBER ---
elif opcao == "💰 Contas a Receber":
    st.title("💰 Contas a Receber")
    st.caption("Gerencie quem te deve e acompanhe os recebimentos previstos.")

    usuario_atual_id = st.session_state.get("usuario_id")

    # EXPANDER: FORMULÁRIO DE CADASTRO
    with st.expander("➕ Cadastrar Nova Conta a Receber"):
        with st.form("form_conta_receber", clear_on_submit=True):
            col_f1, col_f2 = st.columns(2)
            devedor = col_f1.text_input("Nome / Devedor *")
            valor = col_f2.number_input("Valor (R$) *", min_value=0.01, step=10.0, format="%.2f")
            data_prev = st.date_input("Data de Recebimento", value=pd.to_datetime("today"))
            btn_salvar = st.form_submit_button("Salvar Conta a Receber", type="primary")

            if btn_salvar:
                if not devedor.strip():
                    st.error("Preencha o nome do devedor / descrição.")
                else:
                    sucesso = salvar_conta_a_receber(usuario_atual_id, devedor, valor, data_prev)
                    if sucesso:
                        st.success(f"Conta de '{devedor}' salva com sucesso!")
                        st.cache_data.clear()
                        st.rerun()

    # BUSCA REGISTROS NO SUPABASE
    registros = buscar_contas_a_receber(usuario_atual_id)

    if not registros:
        st.info("Nenhuma conta a receber cadastrada.")
    else:
        df_rec = pd.DataFrame(registros)

        if "recebido" not in df_rec.columns:
            df_rec["recebido"] = False
        df_rec["recebido"] = df_rec["recebido"].fillna(False).astype(bool)

        col_data = "data_recebimento" if "data_recebimento" in df_rec.columns else "data"

        # CÁLCULO DAS MÉTRICAS (KPIS)
        df_pendentes = df_rec[df_rec["recebido"] == False]
        df_recebidos = df_rec[df_rec["recebido"] == True]

        val_pendente = df_pendentes["valor"].sum() if not df_pendentes.empty else 0.0
        val_recebido = df_recebidos["valor"].sum() if not df_recebidos.empty else 0.0
        val_total = df_rec["valor"].sum()

        col_kpi1, col_kpi2, col_kpi3 = st.columns(3)
        col_kpi1.metric("⏳ Total Pendente", fmt_moeda(val_pendente))
        col_kpi2.metric("✅ Total Recebido", fmt_moeda(val_recebido))
        col_kpi3.metric("📊 Total Geral", fmt_moeda(val_total))

        st.markdown("---")

        # LISTA COM EXPANDER (SETA) PARA CADA ITEM
        for idx, row in df_rec.iterrows():
            rec_id = row["id"]
            devedor_nome = row.get("descricao") or "Sem descrição"
            valor_fmt = fmt_moeda(row["valor"])
            
            dt_raw = row.get(col_data)
            dt_fmt = pd.to_datetime(dt_raw).strftime("%d/%m/%Y") if pd.notna(dt_raw) else "N/A"
            esta_recebido = row["recebido"]

            # Ícone e texto do cabeçalho da seta
            status_icone = "✅" if esta_recebido else "⏳"
            titulo_expander = f"{status_icone} {devedor_nome} — {valor_fmt} | Previsão: {dt_fmt}"

            # SETA/EXPANDER QUE ESCONDE AS AÇÕES
            with st.expander(titulo_expander):
                st.write(f"**Status:** {'✅ Recebido' if esta_recebido else '⏳ Pendente'}")
                st.write(f"**Valor:** {valor_fmt}")
                st.write(f"**Data:** {dt_fmt}")

                col_btn1, col_btn2 = st.columns(2)

                with col_btn1:
                    if esta_recebido:
                        if st.button("↩️ Desfazer", key=f"undo_rec_{idx}_{rec_id}"):
                            alternar_status_contas_a_receber(rec_id, recebido_atual=True)
                            st.cache_data.clear()
                            st.rerun()
                    else:
                        if st.button("✅ Receber", key=f"btn_rec_{idx}_{rec_id}", type="primary"):
                            alternar_status_contas_a_receber(rec_id, recebido_atual=False)
                            st.cache_data.clear()
                            st.rerun()

                with col_btn2:
                    if st.button("🗑️ Excluir", key=f"del_rec_{idx}_{rec_id}"):
                        excluir_conta_a_receber(usuario_atual_id, rec_id)
                        st.cache_data.clear()
                        st.rerun()